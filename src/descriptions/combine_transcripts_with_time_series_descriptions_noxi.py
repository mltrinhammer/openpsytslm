"""
Combine time-series descriptions (AU patterns) with transcript summaries (speech content)
for NoXi sessions using Gemma 7B-it to describe associations.

This script:
1. Discovers NoXi sessions from data_dir, filtered by NoXi_MetaData.xlsx language
2. Loads time-series descriptions from session_XXX_descriptions.json (output of generate_time_series_descriptions_noxi.py)
3. Loads transcript summaries from session_XXX.summary.json (output of summarize_noxi.py)
4. Matches entries by turn_index and time window (start_ms, end_ms)
5. Uses Gemma 7B-it to combine AU patterns and speech content into coherent paragraphs
6. Saves combined results to output JSON files
"""

import sys
import os
import torch
import torch.multiprocessing as mp
import json
import argparse
import numpy as np
import openpyxl
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from tqdm import tqdm
from transformers import AutoTokenizer, AutoModelForCausalLM

# GPU monitoring
try:
    from pynvml import nvmlInit, nvmlDeviceGetHandleByIndex, nvmlDeviceGetMemoryInfo, nvmlDeviceGetUtilizationRates, nvmlDeviceGetTemperature, NVML_TEMPERATURE_GPU
    PYNVML_AVAILABLE = True
except ImportError:
    PYNVML_AVAILABLE = False


# ============================================================================
# GPU MONITORING
# ============================================================================

class GPUMonitor:
    """Monitor GPU memory and utilization."""

    def __init__(self, gpu_id: int):
        self.gpu_id = gpu_id
        self.initialized = False

        if PYNVML_AVAILABLE:
            try:
                nvmlInit()
                self.handle = nvmlDeviceGetHandleByIndex(gpu_id)
                self.initialized = True
            except Exception as e:
                print(f"⚠️ [GPU {gpu_id}] nvidia-ml-py init failed: {e}")

    def get_stats(self) -> Dict:
        stats = {
            'memory_used_gb': 0, 'memory_total_gb': 0,
            'memory_percent': 0, 'utilization': 0, 'temperature': 0
        }
        if not self.initialized:
            if torch.cuda.is_available():
                try:
                    stats['memory_used_gb'] = torch.cuda.memory_allocated(self.gpu_id) / 1e9
                    stats['memory_total_gb'] = torch.cuda.get_device_properties(self.gpu_id).total_memory / 1e9
                    stats['memory_percent'] = (stats['memory_used_gb'] / stats['memory_total_gb']) * 100
                except Exception:
                    pass
            return stats
        try:
            mem_info = nvmlDeviceGetMemoryInfo(self.handle)
            util_info = nvmlDeviceGetUtilizationRates(self.handle)
            temp = nvmlDeviceGetTemperature(self.handle, NVML_TEMPERATURE_GPU)
            stats['memory_used_gb'] = mem_info.used / 1e9
            stats['memory_total_gb'] = mem_info.total / 1e9
            stats['memory_percent'] = (mem_info.used / mem_info.total) * 100
            stats['utilization'] = util_info.gpu
            stats['temperature'] = temp
        except Exception:
            pass
        return stats

    def format_memory_bar(self, width: int = 25) -> str:
        stats = self.get_stats()
        filled = int((stats['memory_percent'] / 100) * width)
        bar = '█' * filled + '░' * (width - filled)
        return f"[{bar}] {stats['memory_used_gb']:.1f}/{stats['memory_total_gb']:.1f}GB ({stats['memory_percent']:.1f}%) | Util: {stats['utilization']}%"

    def log_status(self, message: str):
        memory_bar = self.format_memory_bar()
        print(f"🎮 [GPU {self.gpu_id}] {message}")
        print(f"   {memory_bar}")


def get_available_gpus() -> List[int]:
    if not torch.cuda.is_available():
        return []
    visible = os.environ.get('CUDA_VISIBLE_DEVICES', '')
    if visible:
        return list(range(len(visible.split(','))))
    return list(range(torch.cuda.device_count()))


# ============================================================================
# METADATA
# ============================================================================

def load_valid_session_ids(metadata_path: str, allowed_languages: List[str] = None) -> Dict[int, str]:
    """Load session IDs from NoXi_MetaData.xlsx that match allowed languages.

    Returns:
        Dict mapping session_id (int) -> language (str)
    """
    if allowed_languages is None:
        allowed_languages = ["French", "German", "English"]

    allowed_lower = [lang.lower() for lang in allowed_languages]

    wb = openpyxl.load_workbook(metadata_path, data_only=True)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    session_id_col = header.index("session_id")
    language_col = header.index("Language")

    valid_sessions = {}
    for row in ws.iter_rows(min_row=2):
        session_id = row[session_id_col].value
        language = row[language_col].value

        if session_id is not None and language is not None:
            if language.strip().lower() in allowed_lower:
                valid_sessions[int(session_id)] = language.strip()

    wb.close()
    return valid_sessions


# ============================================================================
# DATA LOADING
# ============================================================================

def load_timeseries_descriptions(descriptions_dir: Path) -> Dict[str, List[Dict]]:
    """Load all time-series description JSON files from directory.

    Files are named: session_XXX_descriptions.json
    (output from generate_time_series_descriptions_noxi.py)

    Returns:
        Dict mapping session_id_str (e.g. "026") -> list of description entries
    """
    descriptions_by_session = {}

    print(f"\n📂 Loading time-series descriptions from {descriptions_dir}...")
    json_files = list(descriptions_dir.glob("session_*_descriptions.json"))

    for json_file in json_files:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # Extract session_id from filename: session_026_descriptions.json -> 026
        stem = json_file.stem  # session_026_descriptions
        session_id = stem.replace('session_', '').replace('_descriptions', '')
        descriptions_by_session[session_id] = data

    print(f"✅ Loaded {len(json_files)} description files for {len(descriptions_by_session)} sessions")
    return descriptions_by_session


def load_summaries(summary_dir: Path) -> Dict[str, List[Dict]]:
    """Load all summary JSON files from directory.

    Files are named: session_XXX.summary.json
    (output from summarize_noxi.py)

    Returns:
        Dict mapping session_id_str (e.g. "026") -> list of summary entries
    """
    summaries_by_session = {}

    print(f"\n📂 Loading summaries from {summary_dir}...")
    json_files = list(summary_dir.glob("session_*.summary.json"))

    for json_file in json_files:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # Extract session_id from filename: session_026.summary.json -> 026
        stem = json_file.stem  # session_026.summary
        session_id = stem.replace('session_', '').replace('.summary', '')

        # Extract summaries list
        if isinstance(data, dict) and 'summaries' in data:
            summaries_by_session[session_id] = data['summaries']
        elif isinstance(data, list):
            summaries_by_session[session_id] = data
        else:
            print(f"⚠️ Invalid summary format for {json_file.name}")

    print(f"✅ Loaded {len(json_files)} summary files for {len(summaries_by_session)} sessions")
    return summaries_by_session


# ============================================================================
# MATCHING
# ============================================================================

def match_entries(
    descriptions: List[Dict],
    summaries: List[Dict],
    tolerance_ms: int = 500
) -> List[Tuple[Dict, Dict]]:
    """Match time-series description and summary entries by turn_index and time window.

    Args:
        descriptions: List of time-series description entries
        summaries: List of summary entries from summarize_noxi.py
        tolerance_ms: Time window tolerance in milliseconds for matching

    Returns:
        List of (description, summary) tuples for matched entries
    """
    matches = []

    # Build lookup by turn_index for summaries
    summaries_indexed = {}
    for idx, summ in enumerate(summaries):
        # summarize_noxi.py doesn't use turn_indices; entries are indexed sequentially
        summaries_indexed[idx] = summ

    for desc in descriptions:
        turn_idx = desc['turn_index']
        start_ms = desc['start_ms']
        end_ms = desc['end_ms']

        # Try exact turn index match first
        if turn_idx in summaries_indexed:
            summ = summaries_indexed[turn_idx]
            # Verify time windows overlap
            if (start_ms <= summ['end_ms'] + tolerance_ms and
                end_ms >= summ['start_ms'] - tolerance_ms):
                matches.append((desc, summ))
                continue

        # Fallback: search by time window overlap
        for summ in summaries:
            if (start_ms <= summ['end_ms'] + tolerance_ms and
                end_ms >= summ['start_ms'] - tolerance_ms):
                matches.append((desc, summ))
                break

    return matches


# ============================================================================
# COMBINATION PROMPT & LLM
# ============================================================================

def create_combination_prompt(
    timeseries_description: str,
    summary: str,
    speaker_id: str
) -> str:
    """Create prompt for Gemma to combine time-series description and summary."""

    # Map speaker_id to readable label
    speaker_label = speaker_id.capitalize()  # "expert" -> "Expert", "novice" -> "Novice"

    prompt = f"""
You are describing the content in a speech turn from a dyadic interaction. 
Your task is to combine what was said with the speakers' facial expressions to one short, coherent paragraph.

Data for this turn:

Speech content summary: {summary} (spoken by {speaker_label})

Facial Action Unit (AU) patterns: {timeseries_description}

Instructions:
- Begin by describing the speech content very briefly
- Then briefly note any salient facial Action Units (AUs) that stand out for both novice and expert, while only mentioning the most relevant ones.
- Do **not** over-analyze or speculate; be very true to what is actually present in the data available. 
- Do not reflect on the emotional bond, synchrony or similar aspects of the interaction.
- Write your description as a single, natural paragraph — do not use bullet points, numbered steps, or section headings.
- ONLY output your description, do not make any introductory comments.

Description:"""

    return prompt


def combine_with_gemma(
    tokenizer,
    model,
    device: str,
    timeseries_description: str,
    summary: str,
    speaker_id: str,
    use_concat: bool = False
) -> str:
    """Use Gemma 7B-it to combine time-series description and summary into coherent text."""

    if use_concat:
        return f"{summary} {timeseries_description}"

    prompt = create_combination_prompt(
        timeseries_description, summary, speaker_id
    )

    messages = [{"role": "user", "content": prompt}]

    if hasattr(tokenizer, "apply_chat_template"):
        formatted_prompt = tokenizer.apply_chat_template(
            messages,
            tokenize=False,
            add_generation_prompt=True
        )
    else:
        formatted_prompt = prompt

    inputs = tokenizer(formatted_prompt, return_tensors="pt").to(device)

    with torch.inference_mode():
        outputs = model.generate(
            **inputs,
            max_new_tokens=350,
            do_sample=False,
            temperature=None,
            pad_token_id=tokenizer.pad_token_id,
            eos_token_id=tokenizer.eos_token_id
        )

    generated_tokens = outputs[0][inputs['input_ids'].shape[1]:]
    combined = tokenizer.decode(generated_tokens, skip_special_tokens=True).strip()

    return combined


# ============================================================================
# SESSION PROCESSING
# ============================================================================

def process_session(
    session_id: str,
    descriptions: List[Dict],
    summaries: List[Dict],
    tokenizer,
    model,
    device: str,
    max_turns: Optional[int] = None,
    use_concat: bool = False
) -> List[Dict[str, Any]]:
    """Process all turns for a single NoXi session.

    Args:
        session_id: Session identifier (e.g. "026")
        descriptions: List of time-series description entries
        summaries: List of summary entries
        tokenizer: Gemma tokenizer
        model: Gemma model
        device: Device to run on
        max_turns: Maximum number of turns to process (None = all)
        use_concat: If True, use simple concatenation instead of LLM
    """
    results = []

    # Match time-series descriptions with summaries
    matches = match_entries(descriptions, summaries)

    if not matches:
        print(f"⚠️ No matches found for session {session_id}")
        return results

    if max_turns is not None and len(matches) > max_turns:
        matches = matches[:max_turns]
        print(f"ℹ️  Limited to first {max_turns} turns for debugging")

    print(f"\nProcessing session {session_id}: {len(matches)} matched turns")

    skipped_empty = 0

    for desc, summ in tqdm(matches, desc=f"Session {session_id}"):
        try:
            # Skip if time-series description is empty or missing
            description_text = desc.get('generated_description', '').strip()
            if not description_text or description_text.lower().startswith('error:'):
                skipped_empty += 1
                continue

            # Skip if summary is empty
            summary_text = summ.get('summary', summ.get('original_text', summ.get('text', ''))).strip()
            if not summary_text:
                skipped_empty += 1
                continue

            combined = combine_with_gemma(
                tokenizer,
                model,
                device,
                description_text,
                summary_text,
                desc['speaker_id'],
                use_concat=use_concat
            )

            result = {
                "session_id": session_id,
                "turn_index": desc['turn_index'],
                "speaker_id": desc['speaker_id'],
                "start_ms": desc['start_ms'],
                "end_ms": desc['end_ms'],
                "duration_ms": desc.get('duration_ms', desc['end_ms'] - desc['start_ms']),
                "original_timeseries_description": description_text,
                "original_summary": summary_text,
                "combined_description": combined
            }
            results.append(result)

        except Exception as e:
            print(f"❌ Error processing turn {desc['turn_index']}: {e}")
            continue

    if skipped_empty > 0:
        print(f"⚠️  Skipped {skipped_empty} turn(s) with empty description or summary")

    return results


def save_results(results: List[Dict[str, Any]], output_path: Path):
    """Save combined results to JSON."""
    print(f"\n💾 Saving {len(results)} results to {output_path}...")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"✅ Results saved")


# ============================================================================
# MULTI-GPU WORKER
# ============================================================================

def gpu_worker(
    gpu_id: int,
    session_assignments: List[str],
    descriptions_by_session: Dict[str, List[Dict]],
    summaries_by_session: Dict[str, List[Dict]],
    output_dir: Path,
    model_name: str,
    max_turns: Optional[int],
    use_concat: bool,
    skip_existing: bool,
    result_queue: mp.Queue
):
    """Worker function that runs on a specific GPU."""
    try:
        device = f"cuda:{gpu_id}"
        torch.cuda.set_device(gpu_id)

        print(f"\n🚀 [GPU {gpu_id}] Worker starting with {len(session_assignments)} sessions")

        monitor = GPUMonitor(gpu_id)
        monitor.log_status("Worker starting")

        # Load model on this GPU
        if not use_concat:
            print(f"🔧 [GPU {gpu_id}] Loading {model_name}...")
            tokenizer = AutoTokenizer.from_pretrained(model_name)
            model = AutoModelForCausalLM.from_pretrained(
                model_name,
                torch_dtype=torch.float16,
                device_map=device
            )
            model.eval()
            monitor.log_status("Model loaded")
        else:
            tokenizer = None
            model = None
            print(f"⚡ [GPU {gpu_id}] Using concat mode - no model loaded")

        # Process assigned sessions
        worker_results = []
        total_turns_processed = 0

        for session_idx, session_id in enumerate(session_assignments):
            print(f"\n{'='*60}")
            print(f"[GPU {gpu_id}] Session {session_idx + 1}/{len(session_assignments)}: {session_id}")
            print(f"{'='*60}")

            output_file = output_dir / f"session_{session_id}_combined.json"
            if skip_existing and output_file.exists():
                print(f"[GPU {gpu_id}] ⏭️  Skipping session {session_id} - output already exists")
                continue

            if session_id not in descriptions_by_session:
                print(f"[GPU {gpu_id}] ⚠️ No AU descriptions for session {session_id}, skipping")
                continue

            if session_id not in summaries_by_session:
                print(f"[GPU {gpu_id}] ⚠️ No summaries for session {session_id}, skipping")
                continue

            results = process_session(
                session_id,
                descriptions_by_session[session_id],
                summaries_by_session[session_id],
                tokenizer,
                model,
                device,
                max_turns=max_turns,
                use_concat=use_concat
            )

            if results:
                save_results(results, output_file)
                worker_results.append({
                    'session_id': session_id,
                    'num_turns': len(results),
                    'output_path': str(output_file)
                })
                total_turns_processed += len(results)

            if (session_idx + 1) % 3 == 0 or session_idx == len(session_assignments) - 1:
                monitor.log_status(f"Progress: {session_idx + 1}/{len(session_assignments)} sessions")

        # Clean up
        if model is not None:
            del model
            del tokenizer
        torch.cuda.empty_cache()

        result_queue.put({
            'gpu_id': gpu_id,
            'status': 'success',
            'results': worker_results,
            'total_turns': total_turns_processed
        })

        print(f"\n✅ [GPU {gpu_id}] Worker completed: {total_turns_processed} turns across {len(worker_results)} sessions")

    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(f"\n❌ [GPU {gpu_id}] Worker failed: {e}")
        print(error_msg)
        result_queue.put({
            'gpu_id': gpu_id,
            'status': 'error',
            'error': str(e),
            'traceback': error_msg
        })


def distribute_sessions(session_ids: List[str], num_gpus: int) -> Dict[int, List[str]]:
    """Distribute sessions across GPUs using round-robin assignment."""
    assignments = {gpu_id: [] for gpu_id in range(num_gpus)}

    for idx, session_id in enumerate(session_ids):
        gpu_id = idx % num_gpus
        assignments[gpu_id].append(session_id)

    print("\n📋 Session distribution across GPUs:")
    for gpu_id, sessions in assignments.items():
        print(f"  GPU {gpu_id}: {len(sessions)} sessions")

    return assignments


# ============================================================================
# MAIN
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Combine time-series descriptions with transcript summaries for NoXi sessions using Gemma 7B-it"
    )
    parser.add_argument(
        "--descriptions_dir", type=Path, required=True,
        help="Directory containing session_XXX_descriptions.json files (output of generate_time_series_descriptions_noxi.py)"
    )
    parser.add_argument(
        "--summary_dir", type=Path, required=True,
        help="Directory containing session_XXX.summary.json files (output of summarize_noxi.py)"
    )
    parser.add_argument(
        "--metadata", type=str, required=True,
        help="Path to NoXi_MetaData.xlsx"
    )
    parser.add_argument(
        "--output_dir", type=Path, required=True,
        help="Output directory for combined results"
    )
    parser.add_argument(
        "--model_name", type=str, default="google/gemma-7b-it",
        help="Gemma model to use (default: google/gemma-7b-it)"
    )
    parser.add_argument(
        "--max_sessions", type=int, default=None,
        help="Maximum number of sessions to process (None = all)"
    )
    parser.add_argument(
        "--max_turns", type=int, default=None,
        help="Maximum number of speech turns to process per session (None = all, useful for debugging)"
    )
    parser.add_argument(
        "--concat", action="store_true",
        help="Use simple string concatenation instead of LLM (bypasses Gemma model)"
    )
    parser.add_argument(
        "--skip_existing", action="store_true",
        help="Skip sessions that already have output JSON files"
    )
    parser.add_argument(
        "--gpus", type=str, default="0,1,2,3",
        help="Comma-separated GPU indices to use (default: 0,1,2,3)"
    )
    parser.add_argument(
        "--single_gpu", action="store_true",
        help="Force single-GPU mode (no multiprocessing)"
    )
    parser.add_argument(
        "--languages", type=str, default="French,German,English",
        help="Comma-separated list of languages to include (default: French,German,English)"
    )

    args = parser.parse_args()

    # Determine GPUs to use
    if args.gpus:
        gpu_ids = [int(x.strip()) for x in args.gpus.split(',')]
    else:
        gpu_ids = get_available_gpus()

    if not gpu_ids and not args.concat:
        print("❌ No GPUs available! Use --concat for CPU-only mode")
        sys.exit(1)

    num_gpus = len(gpu_ids) if not args.single_gpu else min(1, len(gpu_ids))

    print("🚀 Starting NoXi time-series description + summary combination" + (" (concat mode)" if args.concat else " with Gemma 7B-it"))
    print("=" * 80)
    print(f"Configuration:")
    print(f"  Descriptions dir: {args.descriptions_dir}")
    print(f"  Summary dir: {args.summary_dir}")
    print(f"  Metadata: {args.metadata}")
    print(f"  Output dir: {args.output_dir}")
    print(f"  Mode: {'String concatenation' if args.concat else f'LLM ({args.model_name})'}")
    print(f"  GPUs: {gpu_ids if gpu_ids else 'CPU only'}")
    print(f"  Multi-GPU mode: {'No (single-GPU)' if args.single_gpu or num_gpus <= 1 else f'Yes ({num_gpus} GPUs)'}")
    print(f"  Skip existing: {args.skip_existing}")
    if args.max_turns:
        print(f"  Max turns per session: {args.max_turns} (debugging mode)")
    print()

    # Setup
    args.output_dir.mkdir(parents=True, exist_ok=True)

    # Load valid session IDs from metadata
    allowed_languages = [lang.strip() for lang in args.languages.split(",")]
    valid_sessions = load_valid_session_ids(args.metadata, allowed_languages)
    print(f"📋 Found {len(valid_sessions)} sessions matching languages: {allowed_languages}")

    # Load data
    descriptions_by_session = load_timeseries_descriptions(args.descriptions_dir)
    summaries_by_session = load_summaries(args.summary_dir)

    # Filter to sessions that are valid (language filter) and have both descriptions and summaries
    session_ids = []
    for session_id_str, desc_list in descriptions_by_session.items():
        sid_int = int(session_id_str)
        if sid_int not in valid_sessions:
            continue
        if session_id_str not in summaries_by_session:
            print(f"⚠️ Session {session_id_str}: has descriptions but no summaries, skipping")
            continue
        session_ids.append(session_id_str)

    session_ids.sort()

    if args.max_sessions:
        session_ids = session_ids[:args.max_sessions]

    if not session_ids:
        print("No valid sessions found to process.")
        return

    print(f"\n📊 Processing {len(session_ids)} sessions")

    # =========================================================================
    # SINGLE-GPU MODE (or concat mode)
    # =========================================================================
    if args.single_gpu or num_gpus <= 1 or args.concat:
        device = f"cuda:{gpu_ids[0]}" if gpu_ids and not args.concat else "cpu"
        print(f"\n🔧 Using device: {device}")

        if gpu_ids and not args.concat:
            torch.cuda.set_device(gpu_ids[0])

        if args.concat:
            print("\n⚡ Using simple concatenation mode - skipping model load")
            tokenizer = None
            model = None
        else:
            print(f"\n🔧 Loading {args.model_name}...")
            tokenizer = AutoTokenizer.from_pretrained(args.model_name)
            model = AutoModelForCausalLM.from_pretrained(
                args.model_name,
                torch_dtype=torch.float16 if device != "cpu" else torch.float32,
                device_map=device if device != "cpu" else None,
                low_cpu_mem_usage=True
            )
            model.eval()

            print(f"✅ Model loaded on {device}")
            if device != "cpu":
                print(f"   GPU Memory allocated: {torch.cuda.memory_allocated(gpu_ids[0]) / 1024**3:.2f} GB")

        # Process all sessions
        all_results = []
        processed_count = 0
        skipped_count = 0

        for session_idx, session_id in enumerate(session_ids):
            print(f"\n{'='*80}")
            print(f"Session {session_idx + 1}/{len(session_ids)}: {session_id}")
            print(f"{'='*80}")

            output_file = args.output_dir / f"session_{session_id}_combined.json"
            if args.skip_existing and output_file.exists():
                print(f"⏭️  Skipping session {session_id} - output already exists")
                skipped_count += 1
                continue

            if session_id not in descriptions_by_session:
                print(f"⚠️ No time-series descriptions found for session {session_id}, skipping")
                continue

            if session_id not in summaries_by_session:
                print(f"⚠️ No summaries found for session {session_id}, skipping")
                continue

            results = process_session(
                session_id,
                descriptions_by_session[session_id],
                summaries_by_session[session_id],
                tokenizer,
                model,
                device,
                max_turns=args.max_turns,
                use_concat=args.concat
            )

            if results:
                all_results.extend(results)
                processed_count += 1
                save_results(results, output_file)

        print(f"\n✅ Complete! Combined {len(all_results)} turns across {processed_count} sessions")
        if skipped_count > 0:
            print(f"⏭️  Skipped {skipped_count} sessions (already processed)")
        print(f"📁 Results saved to: {args.output_dir}")

    # =========================================================================
    # MULTI-GPU MODE
    # =========================================================================
    else:
        print(f"\n🚀 Launching {num_gpus} GPU workers...")

        try:
            mp.set_start_method('spawn', force=True)
        except RuntimeError:
            pass

        # Distribute sessions across GPUs
        session_assignments = distribute_sessions(session_ids, num_gpus)

        result_queue = mp.Queue()

        processes = []
        for gpu_id in gpu_ids[:num_gpus]:
            p = mp.Process(
                target=gpu_worker,
                args=(
                    gpu_id,
                    session_assignments[gpu_id],
                    descriptions_by_session,
                    summaries_by_session,
                    args.output_dir,
                    args.model_name,
                    args.max_turns,
                    args.concat,
                    args.skip_existing,
                    result_queue
                )
            )
            p.start()
            processes.append(p)
            print(f"  Started worker on GPU {gpu_id} (PID: {p.pid})")

        print(f"\n⏳ Waiting for {len(processes)} workers to complete...")

        all_worker_results = []
        for _ in range(len(processes)):
            result = result_queue.get()
            all_worker_results.append(result)
            if result['status'] == 'success':
                print(f"  ✅ GPU {result['gpu_id']} completed: {result['total_turns']} turns")
            else:
                print(f"  ❌ GPU {result['gpu_id']} failed: {result['error']}")

        for p in processes:
            p.join()

        # Summary
        total_turns = sum(r.get('total_turns', 0) for r in all_worker_results if r['status'] == 'success')
        total_sessions = sum(len(r.get('results', [])) for r in all_worker_results if r['status'] == 'success')
        failed_gpus = [r['gpu_id'] for r in all_worker_results if r['status'] == 'error']

        print(f"\n{'='*80}")
        print("🎉 MULTI-GPU PROCESSING COMPLETE")
        print(f"{'='*80}")
        print(f"  Total sessions processed: {total_sessions}")
        print(f"  Total turns processed: {total_turns}")
        print(f"  GPUs used: {num_gpus}")
        if failed_gpus:
            print(f"  ⚠️ Failed GPUs: {failed_gpus}")
        print(f"  📁 Results saved to: {args.output_dir}")


if __name__ == "__main__":
    main()
