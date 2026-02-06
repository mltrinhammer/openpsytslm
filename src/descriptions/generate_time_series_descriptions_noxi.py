"""
This script generates time-series descriptions of the AU for NoXi sessions.
In this step, we give Gemma-3 a heatmap of action unit time series, and simply ask it to describe it.
Its akin to what the opentslm authors did with gpt-4o.

Input data comes from extract_aus_from_stream_files.py (expert_aus.csv / novice_aus.csv)
and summaries from summarize_noxi.py (session_XXX.summary.json).

Multi-GPU support for H100 clusters with GPU memory/utilization logging.
"""

import sys
import os
import time
import threading
import torch
import torch.multiprocessing as mp
import numpy as np
import pandas as pd
import json
import argparse
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional
from tqdm import tqdm
import matplotlib.pyplot as plt
from transformers import pipeline
import openpyxl

# GPU monitoring (nvidia-ml-py provides pynvml module)
try:
    import pynvml
    PYNVML_AVAILABLE = True
except ImportError:
    PYNVML_AVAILABLE = False


# ============================================================================
# GPU MONITORING UTILITIES
# ============================================================================

class GPUMonitor:
    """Monitor GPU memory and utilization with nice logging."""

    def __init__(self, gpu_ids: List[int]):
        self.gpu_ids = gpu_ids
        self.initialized = False
        self._lock = threading.Lock()

        if PYNVML_AVAILABLE:
            try:
                pynvml.nvmlInit()
                self.handles = {gpu_id: pynvml.nvmlDeviceGetHandleByIndex(gpu_id)
                               for gpu_id in gpu_ids}
                self.initialized = True
            except Exception as e:
                print(f"⚠️ pynvml initialization failed: {e}")
        else:
            print("⚠️ pynvml not available. Install with: pip install pynvml")

    def get_gpu_stats(self, gpu_id: int) -> Dict[str, Any]:
        """Get memory and utilization stats for a GPU."""
        stats = {
            'gpu_id': gpu_id,
            'memory_used_gb': 0,
            'memory_total_gb': 0,
            'memory_percent': 0,
            'utilization': 0,
            'temperature': 0
        }

        if not self.initialized or gpu_id not in self.handles:
            if torch.cuda.is_available():
                try:
                    stats['memory_used_gb'] = torch.cuda.memory_allocated(gpu_id) / 1e9
                    stats['memory_total_gb'] = torch.cuda.get_device_properties(gpu_id).total_memory / 1e9
                    stats['memory_percent'] = (stats['memory_used_gb'] / stats['memory_total_gb']) * 100
                except Exception:
                    pass
            return stats

        try:
            with self._lock:
                handle = self.handles[gpu_id]
                mem_info = pynvml.nvmlDeviceGetMemoryInfo(handle)
                util_info = pynvml.nvmlDeviceGetUtilizationRates(handle)
                temp = pynvml.nvmlDeviceGetTemperature(handle, pynvml.NVML_TEMPERATURE_GPU)

                stats['memory_used_gb'] = mem_info.used / 1e9
                stats['memory_total_gb'] = mem_info.total / 1e9
                stats['memory_percent'] = (mem_info.used / mem_info.total) * 100
                stats['utilization'] = util_info.gpu
                stats['temperature'] = temp
        except Exception:
            pass

        return stats

    def format_memory_bar(self, gpu_id: int, width: int = 30) -> str:
        """Create a visual memory bar for a GPU."""
        stats = self.get_gpu_stats(gpu_id)
        filled = int((stats['memory_percent'] / 100) * width)
        bar = '█' * filled + '░' * (width - filled)
        return f"GPU {gpu_id}: [{bar}] {stats['memory_used_gb']:.1f}/{stats['memory_total_gb']:.1f}GB ({stats['memory_percent']:.1f}%) | Util: {stats['utilization']}% | Temp: {stats['temperature']}°C"

    def log_all_gpus(self, prefix: str = ""):
        """Log stats for all GPUs."""
        print(f"\n{'='*80}")
        if prefix:
            print(f"📊 {prefix}")
        for gpu_id in self.gpu_ids:
            print(f"  {self.format_memory_bar(gpu_id)}")
        print(f"{'='*80}\n")

    def shutdown(self):
        """Clean up pynvml."""
        if self.initialized:
            try:
                pynvml.nvmlShutdown()
            except Exception:
                pass


def get_available_gpus() -> List[int]:
    """Get list of available GPU indices."""
    if not torch.cuda.is_available():
        return []

    visible = os.environ.get('CUDA_VISIBLE_DEVICES', '')
    if visible:
        return list(range(len(visible.split(','))))

    return list(range(torch.cuda.device_count()))


# ============================================================================
# METADATA
# ============================================================================

def load_valid_session_ids(metadata_path: str, allowed_languages: List[str] = None) -> Dict[int, str]:
    """Load session IDs from NoXi_MetaData.xlsx that match allowed languages.

    Returns:
        Dict mapping session_id (int) -> language (str)
    """
    if allowed_languages is None:
        allowed_languages = ["French", "German", "English"]

    allowed_lower = [lang.lower() for lang in allowed_languages]

    wb = openpyxl.load_workbook(metadata_path, data_only=True)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    session_id_col = header.index("session_id")
    language_col = header.index("Language")

    valid_sessions = {}
    for row in ws.iter_rows(min_row=2):
        session_id = row[session_id_col].value
        language = row[language_col].value

        if session_id is not None and language is not None:
            if language.strip().lower() in allowed_lower:
                valid_sessions[int(session_id)] = language.strip()

    wb.close()
    return valid_sessions


# ============================================================================
# SESSION DISCOVERY
# ============================================================================

def discover_sessions(
    data_dir: Path,
    summary_dir: Path,
    valid_sessions: Dict[int, str],
) -> List[Dict[str, Any]]:
    """Discover NoXi sessions that have AU CSVs and summary JSONs.

    Returns:
        List of dicts with keys: session_id_str, language, expert_au_csv,
        novice_au_csv, summary_path
    """
    sessions = []

    for item in sorted(os.listdir(data_dir)):
        item_path = data_dir / item
        if not item_path.is_dir() or not item.isdigit():
            continue

        session_id = int(item)
        if session_id not in valid_sessions:
            continue

        language = valid_sessions[session_id]

        # Check for AU CSVs
        expert_au_csv = item_path / "expert_aus.csv"
        novice_au_csv = item_path / "novice_aus.csv"

        if not expert_au_csv.exists() or not novice_au_csv.exists():
            print(f"⚠️ Session {item}: AU CSV files not found, skipping")
            continue

        # Check for summary JSON
        summary_path = summary_dir / f"session_{item}.summary.json"
        if not summary_path.exists():
            print(f"⚠️ Session {item}: summary JSON not found at {summary_path}, skipping")
            continue

        sessions.append({
            "session_id_str": item,
            "language": language,
            "expert_au_csv": expert_au_csv,
            "novice_au_csv": novice_au_csv,
            "summary_path": summary_path,
        })

    return sessions


# ============================================================================
# SUMMARY / SPEECH TURN LOADING
# ============================================================================

def load_speech_turns(summary_path: Path) -> List[Dict]:
    """Load speech turns from a summarize_noxi.py output JSON.

    Expected format:
    {
        "session_id": "026",
        "language": "English",
        "summaries": [
            {"speaker_id": "expert", "start_ms": ..., "end_ms": ..., "summary": ..., "original_text": ...},
            ...
        ]
    }
    """
    with open(summary_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    if isinstance(data, dict) and 'summaries' in data:
        return data['summaries']
    elif isinstance(data, list):
        return data
    else:
        return []


def convert_summary_to_turns(summary_data: List[Dict]) -> List[Dict]:
    """Convert summary JSON entries to turn format with turn_index.

    Returns:
        List of turns with turn_index, speaker_id, start_ms, end_ms, text
    """
    turns = []
    for idx, segment in enumerate(summary_data):
        speaker_id = segment.get("speaker_id", "unknown")

        if "start_ms" in segment:
            start_ms = segment["start_ms"]
            end_ms = segment["end_ms"]
        else:
            start_ms = segment.get("start", 0) * 1000
            end_ms = segment.get("end", 0) * 1000

        text = segment.get("summary") or segment.get("original_text") or segment.get("text", "")

        turns.append({
            "turn_index": idx,
            "speaker_id": speaker_id,
            "start_ms": start_ms,
            "end_ms": end_ms,
            "text": text,
        })
    return turns


# ============================================================================
# AU DATA EXTRACTION
# ============================================================================

def extract_au_window(csv_path: Path, start_ms: float, end_ms: float, au_columns: List[str]) -> pd.DataFrame:
    """Extract AU data from an extracted AU CSV for a specific time window.

    The CSV has columns: frame, timestamp (seconds), AU01_r, AU02_r, ...

    Args:
        csv_path: Path to expert_aus.csv or novice_aus.csv
        start_ms: Start time in milliseconds
        end_ms: End time in milliseconds
        au_columns: List of AU column names to extract (e.g. ['AU01_r', 'AU04_r'])

    Returns:
        DataFrame with timestamp_ms and AU columns for the specified window
    """
    df = pd.read_csv(csv_path, skipinitialspace=True)
    df.columns = df.columns.str.strip()

    # Check if requested columns exist
    missing_cols = [col for col in au_columns if col not in df.columns]
    if missing_cols:
        au_cols_in_file = [col for col in df.columns if 'AU' in col.upper()]
        raise KeyError(
            f"Columns {missing_cols} not found in {csv_path.name}. "
            f"Available AU columns: {au_cols_in_file}"
        )

    # Convert timestamp from seconds to milliseconds
    df['timestamp_ms'] = df['timestamp'] * 1000

    # Filter to time window
    mask = (df['timestamp_ms'] >= start_ms) & (df['timestamp_ms'] <= end_ms)
    window_df = df.loc[mask, ['timestamp_ms'] + au_columns].copy()

    return window_df


# ============================================================================
# BINNING AND PLOTTING
# ============================================================================

def bin_time_series(data: pd.DataFrame, au_name: str, num_bins: int = 8) -> np.ndarray:
    """Bin time series data into equal temporal bins and compute mean activation per bin."""
    if len(data) == 0:
        return np.zeros(num_bins)

    bin_indices = np.linspace(0, len(data), num_bins + 1, dtype=int)

    binned_values = []
    for i in range(num_bins):
        start_idx = bin_indices[i]
        end_idx = bin_indices[i + 1]
        if end_idx > start_idx:
            bin_mean = data[au_name].iloc[start_idx:end_idx].mean()
            binned_values.append(bin_mean)
        else:
            binned_values.append(0.0)

    return np.array(binned_values)


def generate_plot_for_turn(
    expert_csv: Path,
    novice_csv: Path,
    turn: Dict,
    au_names: List[str],
    output_path: Path,
    num_bins: int = 8
) -> bool:
    """Generate heatmap visualization with binned AU activations for expert and novice.

    Args:
        expert_csv: Path to expert_aus.csv
        novice_csv: Path to novice_aus.csv
        turn: Speech turn dict with start_ms, end_ms, speaker_id, turn_index
        au_names: List of AU names to plot
        output_path: Where to save the plot
        num_bins: Number of temporal bins (default 8)

    Returns:
        True if successful, False otherwise
    """
    try:
        start_ms = turn['start_ms']
        end_ms = turn['end_ms']
        speaker_id = turn['speaker_id']
        turn_index = turn['turn_index']

        # Extract AU data for both speakers
        expert_data = extract_au_window(expert_csv, start_ms, end_ms, au_names)
        novice_data = extract_au_window(novice_csv, start_ms, end_ms, au_names)

        if expert_data.empty or novice_data.empty:
            print(f"⚠️ No data found for turn {turn_index} ({start_ms}-{end_ms}ms)")
            return False

        # Create binned heatmaps for each AU
        expert_heatmap = np.zeros((len(au_names), num_bins))
        novice_heatmap = np.zeros((len(au_names), num_bins))

        for i, au_name in enumerate(au_names):
            expert_heatmap[i, :] = bin_time_series(expert_data, au_name, num_bins)
            novice_heatmap[i, :] = bin_time_series(novice_data, au_name, num_bins)

        # Create figure with two side-by-side heatmaps
        fig_height = max(6, len(au_names) * 0.5)
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, fig_height))

        # Find global min/max for consistent color scaling
        vmin = min(expert_heatmap.min(), novice_heatmap.min())
        vmax = max(expert_heatmap.max(), novice_heatmap.max())

        # Expert heatmap (left)
        im1 = ax1.imshow(expert_heatmap, aspect='auto', cmap='Blues',
                         interpolation='nearest', vmin=vmin, vmax=vmax)
        ax1.set_title('EXPERT AU Activation', fontsize=14, fontweight='bold', pad=15)
        ax1.set_ylabel('Action Unit', fontsize=12, fontweight='bold')
        ax1.set_xlabel('Time Progression (Start → End)', fontsize=12, fontweight='bold')
        ax1.set_yticks(range(len(au_names)))
        ax1.set_yticklabels(au_names, fontsize=11)
        ax1.set_xticks(range(num_bins))
        ax1.set_xticklabels(['Start', 'Early', 'Early-Mid', 'Mid', 'Late-Mid', 'Late', 'Very Late', 'End'][:num_bins],
                            fontsize=9, rotation=45, ha='right')

        cbar1 = plt.colorbar(im1, ax=ax1, fraction=0.046, pad=0.04)
        cbar1.set_label('Activation Level', fontsize=11, fontweight='bold')

        font_size = 8 if len(au_names) <= 4 else 6
        for i in range(len(au_names)):
            for j in range(num_bins):
                ax1.text(j, i, f'{expert_heatmap[i, j]:.2f}',
                         ha="center", va="center", color="black", fontsize=font_size)

        # Novice heatmap (right)
        im2 = ax2.imshow(novice_heatmap, aspect='auto', cmap='Oranges',
                         interpolation='nearest', vmin=vmin, vmax=vmax)
        ax2.set_title('NOVICE AU Activation', fontsize=14, fontweight='bold', pad=15)
        ax2.set_ylabel('Action Unit', fontsize=12, fontweight='bold')
        ax2.set_xlabel('Time Progression (Start → End)', fontsize=12, fontweight='bold')
        ax2.set_yticks(range(len(au_names)))
        ax2.set_yticklabels(au_names, fontsize=11)
        ax2.set_xticks(range(num_bins))
        ax2.set_xticklabels(['Start', 'Early', 'Early-Mid', 'Mid', 'Mid-Late', 'Late', 'Very Late', 'End'][:num_bins],
                            fontsize=9, rotation=45, ha='right')

        cbar2 = plt.colorbar(im2, ax=ax2, fraction=0.046, pad=0.04)
        cbar2.set_label('Activation Level', fontsize=11, fontweight='bold')

        for i in range(len(au_names)):
            for j in range(num_bins):
                ax2.text(j, i, f'{novice_heatmap[i, j]:.2f}',
                         ha="center", va="center", color="black", fontsize=font_size)

        # Overall title
        plt.suptitle(f"Turn {turn_index}: {speaker_id.capitalize()} speaking ({start_ms:.0f}-{end_ms:.0f}ms)\n" +
                     f"Heatmap shows mean AU activation across {num_bins} temporal phases",
                     fontsize=15, fontweight='bold', y=0.98)

        plt.tight_layout(rect=[0, 0, 1, 0.95])
        plt.savefig(output_path, dpi=150, bbox_inches='tight')
        plt.close()

        return True

    except Exception as e:
        print(f"❌ Error generating plot for turn {turn.get('turn_index', '?')}: {e}")
        return False


# ============================================================================
# DESCRIPTION GENERATION
# ============================================================================

def generate_description_with_pipeline(pipe, image_path: Path, turn: Dict, au_names: List[str]) -> str:
    """Generate time-series description using the Gemma-3 pipeline."""

    pre_prompt = """Describe these AU heatmaps (left=expert blue, right=novice orange) very briefly. 
Each row is one AU across 8 time bins. Write one compact sentence per AU comparing patterns.
ONLY consider the AUs which shows either 1) high variability within expert or novice or 2) strong difference between expert and novice
Consider a maximum of 4 action units. 
Format: "AU##: expert [pattern], novice [pattern], [key difference]."
No markdown, bullets, or headers. Do not refer the facial movement associated with the AU. ONLY output your description
."""

    au_list = ", ".join(au_names)
    context = f"""Turn {turn['turn_index']} ({turn['speaker_id']}), {turn['start_ms']:.0f}-{turn['end_ms']:.0f}ms
AUs: {au_list}
Description:"""

    messages = [
        {
            "role": "user",
            "content": [
                {"type": "image", "url": str(image_path)},
                {"type": "text", "text": f"{pre_prompt}\n\n{context}"}
            ]
        }
    ]

    try:
        output = pipe(
            text=messages,
            max_new_tokens=125,
            do_sample=False
        )
        description = output[0]["generated_text"][-1]["content"].strip()

        # Minimal post-processing
        if description.lower().startswith(("here's", "here is", "the patterns")):
            if ':' in description[:50]:
                description = description.split(':', 1)[1].lstrip()

        import re
        description = description.replace('**', '').replace('*', '')
        description = description.replace('\n', ' ')
        description = re.sub(r'\s+', ' ', description).strip()

        return description
    except Exception as e:
        print(f"❌ Error generating description: {e}")
        return f"Error: {str(e)}"


# ============================================================================
# SESSION PROCESSING
# ============================================================================

def process_session(
    session_info: Dict,
    pipe,
    output_dir: Path,
    au_names: List[str],
    max_turns: int = None,
    skip_existing: bool = True
) -> List[Dict[str, Any]]:
    """Process a single NoXi session for description generation.

    Args:
        session_info: Dict with session_id_str, language, expert_au_csv,
                      novice_au_csv, summary_path
        pipe: Gemma-3 pipeline
        output_dir: Directory to save plots and results
        au_names: List of AU names to analyze
        max_turns: Maximum number of turns to process (None = all)
        skip_existing: If True, skip sessions that already have output JSON files

    Returns:
        List of results dicts (empty if skipped)
    """
    session_id = session_info['session_id_str']
    output_json = output_dir / f"session_{session_id}_descriptions.json"

    if skip_existing and output_json.exists():
        print(f"⏭️ Skipping session {session_id} - output already exists: {output_json}")
        return []

    expert_au_csv = session_info['expert_au_csv']
    novice_au_csv = session_info['novice_au_csv']
    summary_path = session_info['summary_path']

    # Load and convert summary to turns
    summary_data = load_speech_turns(summary_path)
    turns = convert_summary_to_turns(summary_data)

    if max_turns:
        turns = turns[:max_turns]

    results = []

    # Create .temp directory for plots
    temp_dir = output_dir / ".temp"
    temp_dir.mkdir(parents=True, exist_ok=True)

    print(f"\nProcessing session {session_id} [{session_info['language']}]: {len(turns)} turns")

    for turn in tqdm(turns, desc=f"Session {session_id}"):
        turn_index = turn['turn_index']

        # Generate plot in .temp directory
        plot_path = temp_dir / f"session_{session_id}_turn{turn_index:03d}.jpg"
        success = generate_plot_for_turn(expert_au_csv, novice_au_csv, turn, au_names, plot_path)

        if not success:
            continue

        # Generate description
        description = generate_description_with_pipeline(pipe, plot_path, turn, au_names)

        # Collect result
        result = {
            "session_id": session_id,
            "turn_index": turn_index,
            "speaker_id": turn['speaker_id'],
            "start_ms": turn['start_ms'],
            "end_ms": turn['end_ms'],
            "duration_ms": turn['end_ms'] - turn['start_ms'],
            "text": turn['text'],
            "generated_description": description,
            "plot_path": str(plot_path),
            "expert_au_path": str(expert_au_csv),
            "novice_au_path": str(novice_au_csv),
        }
        results.append(result)

    return results


def save_results(results: List[Dict[str, Any]], output_path: Path):
    """Save the generated time-series descriptions to JSON."""
    print(f"\n💾 Saving {len(results)} results to {output_path}...")
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"✅ Results saved")

    if results:
        avg_duration = np.mean([r['duration_ms'] for r in results])
        avg_description_len = np.mean([len(r['generated_description']) for r in results])
        print(f"\n📊 Summary:")
        print(f"  Total turns processed: {len(results)}")
        print(f"  Average turn duration: {avg_duration:.0f}ms")
        print(f"  Average description length: {avg_description_len:.0f} characters")


# ============================================================================
# MULTI-GPU WORKER
# ============================================================================

def gpu_worker(
    gpu_id: int,
    session_assignments: List[Dict],
    output_dir: Path,
    au_columns: List[str],
    max_turns_per_session: Optional[int],
    result_queue: mp.Queue,
    log_interval: int = 10,
    skip_existing: bool = True,
):
    """Worker function that runs on a specific GPU.

    Args:
        gpu_id: GPU index to use
        session_assignments: List of session_info dicts assigned to this GPU
        output_dir: Output directory for results
        au_columns: AU columns to analyze
        max_turns_per_session: Max turns per session (None = all)
        result_queue: Queue to send results back to main process
        log_interval: Log GPU stats every N sessions
        skip_existing: If True, skip sessions that already have output JSON files
    """
    try:
        device = f"cuda:{gpu_id}"
        torch.cuda.set_device(gpu_id)

        print(f"\n🚀 [GPU {gpu_id}] Worker starting with {len(session_assignments)} sessions")

        monitor = GPUMonitor([gpu_id])

        print(f"📊 [GPU {gpu_id}] Memory before model load:")
        print(f"  {monitor.format_memory_bar(gpu_id)}")

        # Load model on this GPU
        start_time = time.time()
        print(f"🔧 [GPU {gpu_id}] Loading Gemma-3-27b-it model...")

        pipe = pipeline(
            "image-text-to-text",
            model="google/gemma-3-27b-it",
            device=device,
            torch_dtype=torch.bfloat16
        )

        load_time = time.time() - start_time
        print(f"✅ [GPU {gpu_id}] Model loaded in {load_time:.1f}s")
        print(f"📊 [GPU {gpu_id}] Memory after model load:")
        print(f"  {monitor.format_memory_bar(gpu_id)}")

        # Process assigned sessions
        worker_results = []
        total_turns_processed = 0

        for session_idx, session_info in enumerate(session_assignments):
            session_id = session_info['session_id_str']

            print(f"\n{'='*60}")
            print(f"[GPU {gpu_id}] Session {session_idx + 1}/{len(session_assignments)}: {session_id}")
            print(f"{'='*60}")

            results = process_session(
                session_info,
                pipe,
                output_dir,
                au_columns,
                max_turns=max_turns_per_session,
                skip_existing=skip_existing,
            )

            # Save immediately after each session
            if results:
                output_json = output_dir / f"session_{session_id}_descriptions.json"
                save_results(results, output_json)
                worker_results.append({
                    'session_id': session_id,
                    'num_turns': len(results),
                    'output_path': str(output_json),
                })
                total_turns_processed += len(results)

            # Log GPU stats periodically
            if (session_idx + 1) % log_interval == 0 or session_idx == len(session_assignments) - 1:
                print(f"\n📊 [GPU {gpu_id}] Progress checkpoint ({session_idx + 1}/{len(session_assignments)} sessions):")
                print(f"  {monitor.format_memory_bar(gpu_id)}")
                print(f"  Total turns processed: {total_turns_processed}")

        # Clean up
        del pipe
        torch.cuda.empty_cache()
        monitor.shutdown()

        result_queue.put({
            'gpu_id': gpu_id,
            'status': 'success',
            'results': worker_results,
            'total_turns': total_turns_processed,
        })

        print(f"\n✅ [GPU {gpu_id}] Worker completed: {total_turns_processed} turns across {len(worker_results)} sessions")

    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(f"\n❌ [GPU {gpu_id}] Worker failed: {e}")
        print(error_msg)
        result_queue.put({
            'gpu_id': gpu_id,
            'status': 'error',
            'error': str(e),
            'traceback': error_msg,
        })


def distribute_sessions(sessions: List[Dict], num_gpus: int) -> Dict[int, List[Dict]]:
    """Distribute sessions across GPUs using round-robin assignment."""
    assignments = {gpu_id: [] for gpu_id in range(num_gpus)}

    for idx, session in enumerate(sessions):
        gpu_id = idx % num_gpus
        assignments[gpu_id].append(session)

    print("\n📋 Session distribution across GPUs:")
    for gpu_id, sess in assignments.items():
        print(f"  GPU {gpu_id}: {len(sess)} sessions")

    return assignments


# ============================================================================
# MAIN
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Generate AU time-series descriptions for NoXi sessions using Gemma-3 multimodal pipeline"
    )
    parser.add_argument(
        "--data_dir", type=Path, required=True,
        help="Root directory containing session subdirectories (026, 027, ...) with AU CSVs"
    )
    parser.add_argument(
        "--summary_dir", type=Path, required=True,
        help="Directory containing session_XXX.summary.json files from summarize_noxi.py"
    )
    parser.add_argument(
        "--metadata", type=str, required=True,
        help="Path to NoXi_MetaData.xlsx"
    )
    parser.add_argument(
        "--output_dir", type=Path, required=True,
        help="Output directory for description JSON files and plots"
    )
    parser.add_argument(
        "--max_sessions", type=int, default=None,
        help="Maximum number of sessions to process (None = all)"
    )
    parser.add_argument(
        "--max_turns_per_session", type=int, default=None,
        help="Maximum turns per session (None = all)"
    )
    parser.add_argument(
        "--au_columns", nargs="+",
        default=['AU01_r', 'AU02_r', 'AU04_r', 'AU05_r'],
        help="AU columns to analyze (default: AU01_r, AU02_r, AU04_r, AU05_r)"
    )
    parser.add_argument(
        "--gpus", type=str, default=None,
        help="Comma-separated GPU indices to use (e.g., '0,1'). Default: all available"
    )
    parser.add_argument(
        "--single_gpu", action="store_true",
        help="Force single-GPU mode (no multiprocessing)"
    )
    parser.add_argument(
        "--log_interval", type=int, default=5,
        help="Log GPU stats every N sessions (default: 5)"
    )
    parser.add_argument(
        "--overwrite", action="store_true",
        help="Overwrite existing output files (default: skip existing)"
    )
    parser.add_argument(
        "--languages", type=str, default="French,German,English",
        help="Comma-separated list of languages to include (default: French,German,English)"
    )

    args = parser.parse_args()

    skip_existing = not args.overwrite

    print("🚀 Starting NoXi AU time-series description generation with Gemma-3 multimodal pipeline")
    print("=" * 80)

    # Determine GPUs to use
    if args.gpus:
        gpu_ids = [int(x.strip()) for x in args.gpus.split(',')]
    else:
        gpu_ids = get_available_gpus()

    if not gpu_ids:
        print("❌ No GPUs available! Falling back to CPU (will be very slow)")
        gpu_ids = []

    num_gpus = len(gpu_ids) if not args.single_gpu else 1

    print(f"\n📊 Configuration:")
    print(f"  Data dir: {args.data_dir}")
    print(f"  Summary dir: {args.summary_dir}")
    print(f"  Metadata: {args.metadata}")
    print(f"  Output dir: {args.output_dir}")
    print(f"  AU columns: {args.au_columns}")
    print(f"  GPUs: {gpu_ids if gpu_ids else 'CPU only'}")
    print(f"  Mode: {'Single-GPU' if args.single_gpu or num_gpus <= 1 else f'Multi-GPU ({num_gpus} GPUs)'}")

    # Create output directory
    args.output_dir.mkdir(parents=True, exist_ok=True)

    # Load valid session IDs from metadata
    allowed_languages = [lang.strip() for lang in args.languages.split(",")]
    valid_sessions = load_valid_session_ids(args.metadata, allowed_languages)
    print(f"\n📋 Found {len(valid_sessions)} sessions matching languages: {allowed_languages}")

    # Discover sessions with AU CSVs and summaries
    sessions = discover_sessions(args.data_dir, args.summary_dir, valid_sessions)

    if args.max_sessions:
        sessions = sessions[:args.max_sessions]

    if not sessions:
        print("No valid sessions found to process.")
        return

    print(f"\n📹 Sessions to process: {len(sessions)}")
    for s in sessions:
        print(f"  Session {s['session_id_str']} ({s['language']})")

    # Initialize GPU monitor
    monitor = GPUMonitor(gpu_ids) if gpu_ids else None

    if monitor:
        monitor.log_all_gpus("Initial GPU Status")

    # =========================================================================
    # SINGLE-GPU MODE (or CPU fallback)
    # =========================================================================
    if args.single_gpu or num_gpus <= 1:
        device = f"cuda:{gpu_ids[0]}" if gpu_ids else "cpu"
        print(f"\n🔧 Using device: {device}")

        if gpu_ids:
            torch.cuda.set_device(gpu_ids[0])

        # Load model
        start_time = time.time()
        print(f"\n🔧 Loading Gemma-3-27b-it model...")

        pipe = pipeline(
            "image-text-to-text",
            model="google/gemma-3-27b-it",
            device=device,
            torch_dtype=torch.bfloat16 if device != "cpu" else torch.float32
        )

        load_time = time.time() - start_time
        print(f"✅ Model loaded in {load_time:.1f}s")

        if monitor:
            monitor.log_all_gpus("After Model Load")

        # Process all sessions
        all_results = []
        saved_files = []

        for session_idx, session_info in enumerate(sessions):
            session_id = session_info['session_id_str']

            print(f"\n{'='*80}")
            print(f"Session {session_idx + 1}/{len(sessions)}: {session_id} [{session_info['language']}]")
            print(f"{'='*80}")

            results = process_session(
                session_info,
                pipe,
                args.output_dir,
                args.au_columns,
                max_turns=args.max_turns_per_session,
                skip_existing=skip_existing,
            )
            all_results.extend(results)

            # Save immediately after each session
            if results:
                output_json = args.output_dir / f"session_{session_id}_descriptions.json"
                save_results(results, output_json)
                saved_files.append(str(output_json))

            # Log GPU stats periodically
            if monitor and ((session_idx + 1) % args.log_interval == 0 or session_idx == len(sessions) - 1):
                monitor.log_all_gpus(f"Progress: {session_idx + 1}/{len(sessions)} sessions")

        print(f"\n✅ Complete! Generated descriptions for {len(all_results)} speech turns across {len(saved_files)} sessions")

    # =========================================================================
    # MULTI-GPU MODE
    # =========================================================================
    else:
        print(f"\n🚀 Launching {num_gpus} GPU workers...")

        try:
            mp.set_start_method('spawn', force=True)
        except RuntimeError:
            pass

        # Distribute sessions across GPUs
        session_assignments = distribute_sessions(sessions, num_gpus)

        # Create result queue
        result_queue = mp.Queue()

        # Launch worker processes
        processes = []
        for gpu_id in gpu_ids[:num_gpus]:
            p = mp.Process(
                target=gpu_worker,
                args=(
                    gpu_id,
                    session_assignments[gpu_id],
                    args.output_dir,
                    args.au_columns,
                    args.max_turns_per_session,
                    result_queue,
                    args.log_interval,
                    skip_existing,
                )
            )
            p.start()
            processes.append(p)
            print(f"  Started worker on GPU {gpu_id} (PID: {p.pid})")

        # Wait for all workers to complete
        print(f"\n⏳ Waiting for {len(processes)} workers to complete...")

        all_worker_results = []
        for _ in range(len(processes)):
            result = result_queue.get()
            all_worker_results.append(result)
            if result['status'] == 'success':
                print(f"  ✅ GPU {result['gpu_id']} completed: {result['total_turns']} turns")
            else:
                print(f"  ❌ GPU {result['gpu_id']} failed: {result['error']}")

        for p in processes:
            p.join()

        # Summary
        total_turns = sum(r.get('total_turns', 0) for r in all_worker_results if r['status'] == 'success')
        total_sessions = sum(len(r.get('results', [])) for r in all_worker_results if r['status'] == 'success')
        failed_gpus = [r['gpu_id'] for r in all_worker_results if r['status'] == 'error']

        print(f"\n{'='*80}")
        print("🎉 MULTI-GPU PROCESSING COMPLETE")
        print(f"{'='*80}")
        print(f"  Total sessions processed: {total_sessions}")
        print(f"  Total turns processed: {total_turns}")
        print(f"  GPUs used: {num_gpus}")
        if failed_gpus:
            print(f"  ⚠️ Failed GPUs: {failed_gpus}")

        if monitor:
            monitor.log_all_gpus("Final GPU Status")

    if monitor:
        monitor.shutdown()


if __name__ == "__main__":
    main()
