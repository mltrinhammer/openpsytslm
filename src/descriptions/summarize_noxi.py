import argparse
import csv
import json
import os
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import torch
import torch.multiprocessing as mp
import openpyxl

# GPU monitoring
try:
    from pynvml import nvmlInit, nvmlDeviceGetHandleByIndex, nvmlDeviceGetMemoryInfo, nvmlDeviceGetUtilizationRates, nvmlDeviceGetTemperature, NVML_TEMPERATURE_GPU
    PYNVML_AVAILABLE = True
except ImportError:
    PYNVML_AVAILABLE = False


# ============================================================================
# GPU MONITORING
# ============================================================================

class GPUMonitor:
    """Monitor GPU memory and utilization."""

    def __init__(self, gpu_id: int):
        self.gpu_id = gpu_id
        self.initialized = False

        if PYNVML_AVAILABLE:
            try:
                nvmlInit()
                self.handle = nvmlDeviceGetHandleByIndex(gpu_id)
                self.initialized = True
            except Exception as e:
                print(f"⚠️ [GPU {gpu_id}] nvidia-ml-py init failed: {e}")

    def get_stats(self) -> Dict:
        """Get GPU memory and utilization stats."""
        stats = {
            'memory_used_gb': 0,
            'memory_total_gb': 0,
            'memory_percent': 0,
            'utilization': 0,
            'temperature': 0
        }

        if not self.initialized:
            if torch.cuda.is_available():
                try:
                    stats['memory_used_gb'] = torch.cuda.memory_allocated(self.gpu_id) / 1e9
                    stats['memory_total_gb'] = torch.cuda.get_device_properties(self.gpu_id).total_memory / 1e9
                    stats['memory_percent'] = (stats['memory_used_gb'] / stats['memory_total_gb']) * 100
                except Exception:
                    pass
            return stats

        try:
            mem_info = nvmlDeviceGetMemoryInfo(self.handle)
            util_info = nvmlDeviceGetUtilizationRates(self.handle)
            temp = nvmlDeviceGetTemperature(self.handle, NVML_TEMPERATURE_GPU)

            stats['memory_used_gb'] = mem_info.used / 1e9
            stats['memory_total_gb'] = mem_info.total / 1e9
            stats['memory_percent'] = (mem_info.used / mem_info.total) * 100
            stats['utilization'] = util_info.gpu
            stats['temperature'] = temp
        except Exception:
            pass

        return stats

    def format_memory_bar(self, width: int = 25) -> str:
        """Create visual memory bar."""
        stats = self.get_stats()
        filled = int((stats['memory_percent'] / 100) * width)
        bar = '█' * filled + '░' * (width - filled)
        return f"[{bar}] {stats['memory_used_gb']:.1f}/{stats['memory_total_gb']:.1f}GB ({stats['memory_percent']:.1f}%) | Util: {stats['utilization']}%"

    def log_status(self, message: str):
        """Log GPU status with memory information."""
        memory_bar = self.format_memory_bar()
        print(f"🎮 [GPU {self.gpu_id}] {message}")
        print(f"   {memory_bar}")


# ============================================================================
# METADATA
# ============================================================================

def load_valid_session_ids(metadata_path: str, allowed_languages: List[str] = None) -> Dict[int, str]:
    """Load session IDs from NoXi_MetaData.xlsx that match allowed languages.

    Returns:
        Dict mapping session_id (int) -> language (str)
    """
    if allowed_languages is None:
        allowed_languages = ["French", "German", "English"]

    allowed_lower = [lang.lower() for lang in allowed_languages]

    wb = openpyxl.load_workbook(metadata_path, data_only=True)
    ws = wb.active

    # Find column indices from header row
    header = [cell.value for cell in ws[1]]
    session_id_col = header.index("session_id")
    language_col = header.index("Language")

    valid_sessions = {}
    for row in ws.iter_rows(min_row=2):
        session_id = row[session_id_col].value
        language = row[language_col].value

        if session_id is not None and language is not None:
            if language.strip().lower() in allowed_lower:
                valid_sessions[int(session_id)] = language.strip()

    wb.close()
    return valid_sessions


# ============================================================================
# TRANSCRIPT LOADING
# ============================================================================

def load_transcript_csv(csv_path: str) -> List[Dict]:
    """Load a NoXi transcript annotation CSV file.

    Format: start_time;end_time;content;confidence
    (No header row, semicolon-separated)

    Returns:
        List of dicts with keys: start, end, text, confidence
    """
    turns = []
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f, delimiter=';')
        for row in reader:
            if len(row) < 3:
                continue
            try:
                start = float(row[0].strip())
                end = float(row[1].strip())
                text = row[2].strip()
                confidence = float(row[3].strip()) if len(row) > 3 else 1.0
                turns.append({
                    "start": start,
                    "end": end,
                    "text": text,
                    "confidence": confidence
                })
            except (ValueError, IndexError):
                continue
    return turns


# ============================================================================
# SPEAKER TURN GROUPING
# ============================================================================

def interleave_and_group_turns(
    expert_turns: List[Dict],
    novice_turns: List[Dict],
    min_duration_s: float = 1.5
) -> List[Dict]:
    """Interleave expert and novice turns by timestamp, then group consecutive
    turns by the same speaker, concatenating their text. Filter out turns
    shorter than min_duration_s.

    Returns:
        List of dicts with keys: speaker_id, text, start_ms, end_ms
    """
    # Tag each turn with speaker
    tagged = []
    for turn in expert_turns:
        tagged.append({**turn, "speaker": "expert"})
    for turn in novice_turns:
        tagged.append({**turn, "speaker": "novice"})

    # Sort by start time
    tagged.sort(key=lambda t: t["start"])

    # Group consecutive same-speaker turns
    grouped = []
    current_speaker = None
    current_texts = []
    current_start = None
    current_end = None

    for turn in tagged:
        text = turn["text"].strip()
        if not text:
            continue

        speaker = turn["speaker"]

        if speaker != current_speaker:
            # Save previous group
            if current_speaker is not None and current_texts:
                duration_s = current_end - current_start
                if duration_s >= min_duration_s:
                    grouped.append({
                        "speaker_id": current_speaker,
                        "text": " ".join(current_texts),
                        "start_ms": int(current_start * 1000),
                        "end_ms": int(current_end * 1000),
                    })

            # Start new group
            current_speaker = speaker
            current_texts = [text]
            current_start = turn["start"]
            current_end = turn["end"]
        else:
            # Same speaker, concatenate
            current_texts.append(text)
            current_end = turn["end"]

    # Don't forget the last group
    if current_speaker is not None and current_texts:
        duration_s = current_end - current_start
        if duration_s >= min_duration_s:
            grouped.append({
                "speaker_id": current_speaker,
                "text": " ".join(current_texts),
                "start_ms": int(current_start * 1000),
                "end_ms": int(current_end * 1000),
            })

    return grouped


# ============================================================================
# SUMMARIZATION PROMPT
# ============================================================================

def create_summary_prompt(speaker_id: str, start_ms: int, end_ms: int, combined_text: str, language: str) -> str:
    """Create prompt for summarizing a speaker turn. Always output in English."""
    start_seconds = max(start_ms, 0) // 1000
    end_seconds = max(end_ms, 0) // 1000
    start_minute, start_second = divmod(start_seconds, 60)
    end_minute, end_second = divmod(end_seconds, 60)
    time_range = f"Time {start_minute:02d}:{start_second:02d}–{end_minute:02d}:{end_second:02d}"

    speaker_label = speaker_id.capitalize()

    language_instruction = ""
    if language.lower() != "english":
        language_instruction = (
            f"The speech excerpt below is in {language}. "
            f"You MUST write the summary in English regardless of the input language.\n"
        )

    return (
        f"You are a concise conversation summarizer. "
        f"Provide a short, anonymous summary (1-2 sentences) of what {speaker_label} said. "
        f"Always write the summary in English.\n"
        f"{language_instruction}"
        f"{speaker_label}'s speech excerpt ({time_range}):\n"
        f"{combined_text}\n"
        f"Summary:"
    )


# ============================================================================
# GPU WORKER
# ============================================================================

def gpu_worker(
    gpu_id: int,
    session_jobs: List[Tuple[str, str, Path, Path]],
    output_dir: Path,
    model_name: str,
    skip_existing: bool,
    min_duration_s: float,
):
    """Worker process for GPU inference.

    session_jobs: list of (session_id_str, language, expert_csv_path, novice_csv_path)
    """
    import torch
    from transformers import AutoTokenizer, AutoModelForCausalLM

    torch.cuda.set_device(gpu_id)

    monitor = GPUMonitor(gpu_id)
    monitor.log_status("Worker starting")

    # Load model
    print(f"[GPU {gpu_id}] Loading model {model_name}...")
    tokenizer = AutoTokenizer.from_pretrained(model_name)
    model = AutoModelForCausalLM.from_pretrained(
        model_name,
        torch_dtype=torch.bfloat16,
        device_map=f"cuda:{gpu_id}"
    )
    model.eval()

    monitor.log_status("Model loaded")

    # Process each session
    for job_idx, (session_id_str, language, expert_csv, novice_csv) in enumerate(session_jobs):
        summary_path = output_dir / f"session_{session_id_str}.summary.json"

        # Skip if exists and flag is set
        if skip_existing and summary_path.exists():
            print(f"[GPU {gpu_id}] Skipping session {session_id_str} (summary exists)")
            continue

        print(f"[GPU {gpu_id}] Processing session {session_id_str} [{language}] ({job_idx + 1}/{len(session_jobs)})")

        try:
            # Load transcripts
            expert_turns = load_transcript_csv(str(expert_csv))
            novice_turns = load_transcript_csv(str(novice_csv))

            if not expert_turns and not novice_turns:
                print(f"[GPU {gpu_id}] No turns found for session {session_id_str}, skipping")
                continue

            # Interleave, group by speaker, and filter short turns
            speaker_turns = interleave_and_group_turns(expert_turns, novice_turns, min_duration_s=min_duration_s)

            if not speaker_turns:
                print(f"[GPU {gpu_id}] No speaker turns (after filtering) for session {session_id_str}, skipping")
                continue

            # Generate summaries
            summaries = []
            for st in speaker_turns:
                prompt = create_summary_prompt(
                    st["speaker_id"],
                    st["start_ms"],
                    st["end_ms"],
                    st["text"],
                    language,
                )

                inputs = tokenizer(prompt, return_tensors="pt").to(f"cuda:{gpu_id}")

                with torch.no_grad():
                    outputs = model.generate(
                        **inputs,
                        max_new_tokens=128,
                        do_sample=False,
                        temperature=None,
                        top_p=None,
                    )

                generated_text = tokenizer.decode(outputs[0], skip_special_tokens=True)
                # Extract just the summary part after "Summary:"
                if "Summary:" in generated_text:
                    summary_text = generated_text.split("Summary:")[-1].strip()
                else:
                    summary_text = generated_text.strip()

                summaries.append({
                    "speaker_id": st["speaker_id"],
                    "start_ms": st["start_ms"],
                    "end_ms": st["end_ms"],
                    "original_text": st["text"],
                    "summary": summary_text,
                })

            # Save summaries
            output_data = {
                "session_id": session_id_str,
                "language": language,
                "transcript_paths": {
                    "expert": str(expert_csv),
                    "novice": str(novice_csv),
                },
                "summaries": summaries,
            }

            with open(summary_path, "w", encoding="utf-8") as f:
                json.dump(output_data, f, indent=2, ensure_ascii=False)

            print(f"[GPU {gpu_id}] Saved {len(summaries)} summaries to {summary_path.name}")

        except Exception as e:
            print(f"[GPU {gpu_id}] Error processing session {session_id_str}: {e}")
            continue

        # Log status every 5 sessions
        if (job_idx + 1) % 5 == 0:
            monitor.log_status(f"Processed {job_idx + 1}/{len(session_jobs)} sessions")

    monitor.log_status("Worker finished")


# ============================================================================
# MAIN
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Generate conversation summaries for NoXi sessions using multi-GPU inference"
    )
    parser.add_argument(
        "--data_dir", type=str, required=True,
        help="Root directory containing session subdirectories (026, 027, ...)"
    )
    parser.add_argument(
        "--metadata", type=str, required=True,
        help="Path to NoXi_MetaData.xlsx"
    )
    parser.add_argument(
        "--output_dir", type=str, required=True,
        help="Directory to save summary JSON files"
    )
    parser.add_argument(
        "--model_name", type=str, default="google/gemma-2-9b-it",
        help="HuggingFace model name (default: google/gemma-2-9b-it)"
    )
    parser.add_argument(
        "--gpus", type=str, default="0",
        help="Comma-separated GPU IDs (default: 0)"
    )
    parser.add_argument(
        "--skip_existing", action="store_true",
        help="Skip sessions that already have summaries"
    )
    parser.add_argument(
        "--min_duration", type=float, default=1.5,
        help="Minimum speaker turn duration in seconds (default: 1.5)"
    )
    parser.add_argument(
        "--languages", type=str, default="French,German,English",
        help="Comma-separated list of languages to include (default: French,German,English)"
    )
    args = parser.parse_args()

    data_dir = Path(args.data_dir)
    if not data_dir.exists():
        raise FileNotFoundError(f"Data directory not found: {data_dir}")

    metadata_path = Path(args.metadata)
    if not metadata_path.exists():
        raise FileNotFoundError(f"Metadata file not found: {metadata_path}")

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Load valid session IDs from metadata
    allowed_languages = [lang.strip() for lang in args.languages.split(",")]
    valid_sessions = load_valid_session_ids(str(metadata_path), allowed_languages)
    print(f"Found {len(valid_sessions)} sessions matching languages: {allowed_languages}")

    # Find session directories with transcript files
    session_jobs = []
    for item in sorted(os.listdir(data_dir)):
        item_path = data_dir / item
        if not item_path.is_dir() or not item.isdigit():
            continue

        session_id = int(item)

        # Check if this session matches our language filter
        if session_id not in valid_sessions:
            continue

        language = valid_sessions[session_id]

        # Look for transcript annotation files
        expert_csv = item_path / "expert.audio.transcript.annotation.csv"
        novice_csv = item_path / "novice.audio.transcript.annotation.csv"

        if not expert_csv.exists() or not novice_csv.exists():
            print(f"⚠️ Session {item}: transcript files not found, skipping")
            continue

        session_jobs.append((item, language, expert_csv, novice_csv))

    if not session_jobs:
        print("No valid sessions found to process.")
        return

    print(f"Found {len(session_jobs)} sessions to process")
    for sid, lang, _, _ in session_jobs:
        print(f"  Session {sid} ({lang})")

    # Parse GPU IDs
    gpu_ids = [int(x.strip()) for x in args.gpus.split(",")]
    num_gpus = len(gpu_ids)
    print(f"\nUsing {num_gpus} GPU(s): {gpu_ids}")

    # Distribute sessions across GPUs (round-robin)
    gpu_assignments = [[] for _ in range(num_gpus)]
    for idx, job in enumerate(session_jobs):
        gpu_idx = idx % num_gpus
        gpu_assignments[gpu_idx].append(job)

    # Print distribution
    for gpu_idx, jobs in enumerate(gpu_assignments):
        print(f"  GPU {gpu_ids[gpu_idx]}: {len(jobs)} sessions")

    # Launch workers
    if num_gpus == 1:
        # Single GPU: run directly (no multiprocessing overhead)
        gpu_worker(gpu_ids[0], gpu_assignments[0], output_dir, args.model_name, args.skip_existing, args.min_duration)
    else:
        mp.set_start_method("spawn", force=True)
        processes = []

        for gpu_idx, gpu_id in enumerate(gpu_ids):
            job_subset = gpu_assignments[gpu_idx]
            if not job_subset:
                continue

            p = mp.Process(
                target=gpu_worker,
                args=(gpu_id, job_subset, output_dir, args.model_name, args.skip_existing, args.min_duration)
            )
            p.start()
            processes.append(p)

        # Wait for all processes
        for p in processes:
            p.join()

    print("\nAll workers finished!")


if __name__ == "__main__":
    main()
