"""
Loader for NoXi dyadic interaction data.

This loader reads:
1. NoXi session directories (results/XXX/) containing:
   - expert_aus.csv / novice_aus.csv — OpenFace AU time series
   - expert.audio.transcript.annotation.csv / novice.audio.transcript.annotation.csv — transcripts
2. Pre-generated combined description JSONs (from combine_transcripts_with_time_series_descriptions_noxi.py)
   OR falls back to building samples from raw AU CSVs + transcript CSVs + summary JSONs
3. NoXi_MetaData.xlsx for language-based session filtering

The loader creates training samples compatible with the opentslm QADataset framework,
mirroring the structure used in har_cot_loader.py and dyadloader.py.
"""

import csv
import json
import os
import hashlib
import pickle
import numpy as np
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datasets import Dataset as HFDataset


# ============================================================================
# CONSTANTS
# ============================================================================

# AU columns to use (intensity-based regression AUs from OpenFace)
DEFAULT_AU_COLUMNS = [
    "AU04_r", "AU06_r", "AU12_r", "AU15_r",
]


# ============================================================================
# METADATA
# ============================================================================

def load_valid_session_ids(
    metadata_path: str,
    allowed_languages: Optional[List[str]] = None
) -> Dict[int, str]:
    """Load session IDs from NoXi_MetaData.xlsx that match allowed languages.

    Returns:
        Dict mapping session_id (int) -> language (str)
    """
    if allowed_languages is None:
        allowed_languages = ["French", "German", "English"]

    allowed_lower = [lang.lower() for lang in allowed_languages]

    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for reading NoXi metadata. Install with: pip install openpyxl")

    wb = openpyxl.load_workbook(metadata_path, data_only=True)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    session_id_col = header.index("session_id")
    language_col = header.index("Language")

    valid_sessions = {}
    for row in ws.iter_rows(min_row=2):
        session_id = row[session_id_col].value
        language = row[language_col].value

        if session_id is not None and language is not None:
            if language.strip().lower() in allowed_lower:
                valid_sessions[int(session_id)] = language.strip()

    wb.close()
    return valid_sessions


# ============================================================================
# TRANSCRIPT LOADING
# ============================================================================

def load_transcript_csv(csv_path: str) -> List[Dict]:
    """Load NoXi transcript annotation CSV (semicolon-separated, no header).

    Format per row: start_time;end_time;text;confidence

    Returns:
        List of dicts with keys: start, end, text, confidence
    """
    turns = []
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f, delimiter=";")
        for row in reader:
            if len(row) < 3:
                continue
            try:
                start = float(row[0].strip())
                end = float(row[1].strip())
                text = row[2].strip()
                confidence = float(row[3].strip()) if len(row) > 3 else 1.0
                turns.append({
                    "start": start,
                    "end": end,
                    "text": text,
                    "confidence": confidence,
                })
            except (ValueError, IndexError):
                continue
    return turns


def interleave_and_group_turns(
    expert_turns: List[Dict],
    novice_turns: List[Dict],
    min_duration_s: float = 1.5,
) -> List[Dict]:
    """Interleave expert and novice turns chronologically, then group consecutive
    turns by the same speaker and filter short turns.

    Returns:
        List of dicts: speaker_id, text, start_ms, end_ms
    """
    tagged = []
    for turn in expert_turns:
        tagged.append({**turn, "speaker": "expert"})
    for turn in novice_turns:
        tagged.append({**turn, "speaker": "novice"})

    tagged.sort(key=lambda t: t["start"])

    grouped = []
    current_speaker = None
    current_texts = []
    current_start = None
    current_end = None

    for turn in tagged:
        text = turn["text"].strip()
        if not text:
            continue

        speaker = turn["speaker"]

        if speaker != current_speaker:
            if current_speaker is not None and current_texts:
                duration_s = (current_end - current_start)
                if duration_s >= min_duration_s:
                    grouped.append({
                        "speaker_id": current_speaker,
                        "text": " ".join(current_texts),
                        "start_ms": int(current_start * 1000),
                        "end_ms": int(current_end * 1000),
                    })
            current_speaker = speaker
            current_texts = [text]
            current_start = turn["start"]
            current_end = turn["end"]
        else:
            current_texts.append(text)
            current_end = turn["end"]

    # Last group
    if current_speaker is not None and current_texts:
        duration_s = (current_end - current_start)
        if duration_s >= min_duration_s:
            grouped.append({
                "speaker_id": current_speaker,
                "text": " ".join(current_texts),
                "start_ms": int(current_start * 1000),
                "end_ms": int(current_end * 1000),
            })

    return grouped


# ============================================================================
# AU DATA EXTRACTION
# ============================================================================

def load_au_csv(csv_path: str) -> pd.DataFrame:
    """Load an OpenFace AU CSV into a DataFrame.

    Args:
        csv_path: Path to *_aus.csv (expert_aus.csv or novice_aus.csv)

    Returns:
        DataFrame with columns: frame, timestamp, AU01_r, ..., AU45_r
    """
    df = pd.read_csv(csv_path)
    df.columns = df.columns.str.strip()
    df["timestamp"] = pd.to_numeric(df["timestamp"], errors="coerce")
    df = df.dropna(subset=["timestamp"])
    return df


def extract_au_window(
    au_df: pd.DataFrame,
    start_s: float,
    end_s: float,
    au_columns: List[str],
    max_seq_length: int = 4096,
) -> Optional[Dict]:
    """Extract AU vectors for a time window.

    Returns:
        Dict with 'au_vectors' (dict of au_col -> list[float]),
                    'au_stats' (dict of au_col -> {mean, std}),
        or None if insufficient data.
    """
    window = au_df[
        (au_df["timestamp"] >= start_s) &
        (au_df["timestamp"] < end_s)
    ]

    if len(window) < 5:
        return None

    au_vectors = {}
    au_stats = {}

    for col in au_columns:
        if col not in window.columns:
            continue
        signal = window[col].to_numpy(dtype=np.float64)

        # Handle NaN from failed face detection frames
        nan_mask = np.isnan(signal)
        if nan_mask.all():
            # All NaN — skip this turn
            return None
        if nan_mask.any():
            # Forward-fill then backward-fill
            df_tmp = pd.Series(signal)
            df_tmp = df_tmp.ffill().bfill()
            signal = df_tmp.to_numpy(dtype=np.float64)

        au_vectors[col] = signal.tolist()
        au_stats[col] = {
            "mean": float(np.nanmean(signal)),
            "std": float(np.nanstd(signal)),
        }

    if not au_vectors:
        return None

    # Downsample if necessary
    sample_len = len(next(iter(au_vectors.values())))
    if sample_len > max_seq_length:
        indices = np.linspace(0, sample_len - 1, max_seq_length, dtype=int)
        for col in au_vectors:
            arr = np.array(au_vectors[col])
            downsampled = arr[indices]
            au_vectors[col] = downsampled.tolist()
            au_stats[col] = {
                "mean": float(downsampled.mean()),
                "std": float(downsampled.std()),
            }

    return {"au_vectors": au_vectors, "au_stats": au_stats}


# ============================================================================
# COMBINED DESCRIPTION LOADING (pre-generated)
# ============================================================================

def load_combined_descriptions(combined_dir: str) -> Dict[str, List[Dict]]:
    """Load pre-generated combined description JSONs.

    Files are named: session_XXX_combined.json
    (output from combine_transcripts_with_time_series_descriptions_noxi.py)

    Returns:
        Dict mapping session_id_str -> list of combined-turn dicts
    """
    combined_by_session = {}
    combined_path = Path(combined_dir)

    if not combined_path.exists():
        print(f"[noxiloader] Combined descriptions directory not found: {combined_dir}")
        return combined_by_session

    for json_file in combined_path.glob("session_*_combined.json"):
        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)
        session_id = json_file.stem.replace("session_", "").replace("_combined", "")
        combined_by_session[session_id] = data

    print(f"[noxiloader] Loaded {len(combined_by_session)} combined description files")
    return combined_by_session


def load_summaries(summary_dir: str) -> Dict[str, List[Dict]]:
    """Load pre-generated summary JSONs.

    Files are named: session_XXX.summary.json
    (output from summarize_noxi.py)

    Returns:
        Dict mapping session_id_str -> list of summary entries
    """
    summaries = {}
    summary_path = Path(summary_dir)

    if not summary_path.exists():
        print(f"[noxiloader] Summary directory not found: {summary_dir}")
        return summaries

    for json_file in summary_path.glob("session_*.summary.json"):
        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)
        session_id = json_file.stem.replace("session_", "").replace(".summary", "")

        if isinstance(data, dict) and "summaries" in data:
            summaries[session_id] = data["summaries"]
        elif isinstance(data, list):
            summaries[session_id] = data

    print(f"[noxiloader] Loaded {len(summaries)} summary files")
    return summaries


# ============================================================================
# SAMPLE BUILDING
# ============================================================================

def _build_samples_for_session(
    session_id: str,
    session_dir: str,
    au_columns: List[str],
    max_seq_length: int,
    combined_entries: Optional[List[Dict]] = None,
    summary_entries: Optional[List[Dict]] = None,
) -> List[Dict]:
    """Build training samples for a single NoXi session.

    If combined_entries are available (from combine_transcripts_with_time_series_descriptions_noxi.py),
    use them as the answer target.
    Otherwise, fall back to raw transcript text as the answer.

    In both cases, we always load the raw AU CSV data for the actual time series vectors.

    Returns:
        List of sample dicts compatible with the QADataset framework.
    """
    samples = []

    expert_au_path = os.path.join(session_dir, "expert_aus.csv")
    novice_au_path = os.path.join(session_dir, "novice_aus.csv")
    expert_transcript_path = os.path.join(session_dir, "expert.audio.transcript.annotation.csv")
    novice_transcript_path = os.path.join(session_dir, "novice.audio.transcript.annotation.csv")

    # Check required files exist
    required_files = [expert_au_path, novice_au_path, expert_transcript_path, novice_transcript_path]
    for fp in required_files:
        if not os.path.exists(fp):
            print(f"[noxiloader] Missing file for session {session_id}: {fp}")
            return samples

    # Load AU CSVs
    expert_au_df = load_au_csv(expert_au_path)
    novice_au_df = load_au_csv(novice_au_path)

    # Filter to available AU columns
    available_au_cols = [c for c in au_columns if c in expert_au_df.columns and c in novice_au_df.columns]
    if not available_au_cols:
        print(f"[noxiloader] No matching AU columns for session {session_id}")
        return samples

    # Strategy 1: Use pre-generated combined descriptions
    if combined_entries is not None and len(combined_entries) > 0:
        for entry in combined_entries:
            start_s = entry["start_ms"] / 1000.0
            end_s = entry["end_ms"] / 1000.0
            speaker_id = entry.get("speaker_id", "unknown")

            expert_data = extract_au_window(expert_au_df, start_s, end_s, available_au_cols, max_seq_length)
            novice_data = extract_au_window(novice_au_df, start_s, end_s, available_au_cols, max_seq_length)

            if expert_data is None or novice_data is None:
                continue

            sample = {
                "session_id": session_id,
                "turn_index": entry.get("turn_index", -1),
                "speaker_id": speaker_id,
                "start_ms": entry["start_ms"],
                "end_ms": entry["end_ms"],
                "window_start": start_s,
                "window_end": end_s,
                "expert_au_vectors": expert_data["au_vectors"],
                "expert_au_stats": expert_data["au_stats"],
                "novice_au_vectors": novice_data["au_vectors"],
                "novice_au_stats": novice_data["au_stats"],
                "au_columns": available_au_cols,
                "original_summary": entry.get("original_summary", ""),
                "original_timeseries_description": entry.get("original_timeseries_description", ""),
                "answer": entry.get("combined_description", ""),
            }
            samples.append(sample)

        return samples

    # Strategy 2: Fall back to building from raw transcripts + optional summaries
    expert_turns = load_transcript_csv(expert_transcript_path)
    novice_turns = load_transcript_csv(novice_transcript_path)
    grouped_turns = interleave_and_group_turns(expert_turns, novice_turns, min_duration_s=1.5)

    # Build a lookup for summaries if available
    summary_lookup = {}
    if summary_entries:
        for idx, summ in enumerate(summary_entries):
            summary_lookup[idx] = summ

    for turn_idx, turn in enumerate(grouped_turns):
        start_s = turn["start_ms"] / 1000.0
        end_s = turn["end_ms"] / 1000.0
        speaker_id = turn["speaker_id"]

        expert_data = extract_au_window(expert_au_df, start_s, end_s, available_au_cols, max_seq_length)
        novice_data = extract_au_window(novice_au_df, start_s, end_s, available_au_cols, max_seq_length)

        if expert_data is None or novice_data is None:
            continue

        # Use summary if available, otherwise use raw text
        original_summary = turn["text"]
        answer_text = turn["text"]  # Fallback: use raw transcript as the answer

        if turn_idx in summary_lookup:
            summ_entry = summary_lookup[turn_idx]
            if isinstance(summ_entry, dict):
                original_summary = summ_entry.get("summary", turn["text"])
                answer_text = original_summary

        sample = {
            "session_id": session_id,
            "turn_index": turn_idx,
            "speaker_id": speaker_id,
            "start_ms": turn["start_ms"],
            "end_ms": turn["end_ms"],
            "window_start": start_s,
            "window_end": end_s,
            "expert_au_vectors": expert_data["au_vectors"],
            "expert_au_stats": expert_data["au_stats"],
            "novice_au_vectors": novice_data["au_vectors"],
            "novice_au_stats": novice_data["au_stats"],
            "au_columns": available_au_cols,
            "original_summary": original_summary,
            "original_timeseries_description": "",
            "answer": answer_text,
        }
        samples.append(sample)

    return samples


# ============================================================================
# SPLIT CREATION
# ============================================================================

def discover_session_dirs(data_dir: str) -> Dict[str, str]:
    """Discover NoXi session directories under data_dir.

    Expected structure: data_dir/XXX/expert_aus.csv
    where XXX is a zero-padded session ID.

    Returns:
        Dict mapping session_id_str -> session_directory_path
    """
    sessions = {}
    data_path = Path(data_dir)

    if not data_path.exists():
        print(f"[noxiloader] Data directory not found: {data_dir}")
        return sessions

    for child in sorted(data_path.iterdir()):
        if not child.is_dir():
            continue
        # Check if this looks like a session directory
        expert_csv = child / "expert_aus.csv"
        novice_csv = child / "novice_aus.csv"
        if expert_csv.exists() and novice_csv.exists():
            sessions[child.name] = str(child)

    return sessions


def load_noxi_cot_splits(
    data_dir: str = "results",
    combined_dir: Optional[str] = None,
    summary_dir: Optional[str] = None,
    metadata_path: Optional[str] = None,
    allowed_languages: Optional[List[str]] = None,
    train_sessions: Optional[List[str]] = None,
    val_sessions: Optional[List[str]] = None,
    test_sessions: Optional[List[str]] = None,
    au_columns: Optional[List[str]] = None,
    max_samples: Optional[int] = None,
    max_seq_length: int = 4096,
) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    """Load NoXi CoT dataset and create train/val/test splits.

    The split is performed at the SESSION level to prevent data leakage
    (all turns from one session stay in the same split).

    Args:
        data_dir: Root directory containing session subdirectories (results/)
        combined_dir: Directory with session_XXX_combined.json files (optional)
        summary_dir: Directory with session_XXX.summary.json files (optional)
        metadata_path: Path to NoXi_MetaData.xlsx for language filtering (optional)
        allowed_languages: Languages to include (default: French, German, English)
        train_sessions: Explicit list of session IDs for training
        val_sessions: Explicit list of session IDs for validation
        test_sessions: Explicit list of session IDs for testing
        au_columns: AU column names to use (default: all 17 intensity AUs)
        max_samples: Max samples per split for debugging
        max_seq_length: Max time series length (longer sequences are downsampled)

    Returns:
        Tuple of (train_samples, val_samples, test_samples)
    """
    if au_columns is None:
        au_columns = DEFAULT_AU_COLUMNS

    # Disk cache
    cache_key = f"{data_dir}_{combined_dir}_{summary_dir}_{metadata_path}_{max_samples}_{au_columns}_{max_seq_length}"
    cache_hash = hashlib.md5(cache_key.encode()).hexdigest()[:8]
    cache_file = Path(f"noxi_cache_{cache_hash}.pkl")

    if cache_file.exists():
        print(f"[noxiloader] Loading from cache: {cache_file}")
        with open(cache_file, "rb") as f:
            return pickle.load(f)

    # Discover sessions
    session_dirs = discover_session_dirs(data_dir)
    print(f"[noxiloader] Discovered {len(session_dirs)} session directories")

    # Filter by language if metadata available
    if metadata_path and os.path.exists(metadata_path):
        valid_sessions = load_valid_session_ids(metadata_path, allowed_languages)
        filtered_dirs = {}
        for sid, sdir in session_dirs.items():
            sid_int = int(sid) if sid.isdigit() else -1
            if sid_int in valid_sessions:
                filtered_dirs[sid] = sdir
        print(f"[noxiloader] After language filter: {len(filtered_dirs)} sessions")
        session_dirs = filtered_dirs

    if not session_dirs:
        print("[noxiloader] No sessions found! Returning empty splits.")
        return [], [], []

    # Load pre-generated data if available
    combined_by_session = {}
    if combined_dir:
        combined_by_session = load_combined_descriptions(combined_dir)

    summaries_by_session = {}
    if summary_dir:
        summaries_by_session = load_summaries(summary_dir)

    # Create splits
    all_session_ids = sorted(session_dirs.keys())

    if train_sessions is None and val_sessions is None and test_sessions is None:
        n_total = len(all_session_ids)
        # Reserve 1 session for test, 5 for validation, rest for training
        n_test = 1
        n_val = 5
        n_train = max(1, n_total - n_val - n_test)

        train_sessions = all_session_ids[:n_train]
        val_sessions = all_session_ids[n_train:n_train + n_val]
        test_sessions = all_session_ids[n_train + n_val:]

    print(f"[noxiloader] Split sizes: train={len(train_sessions)}, val={len(val_sessions)}, test={len(test_sessions)}")

    def _process_split(session_ids: List[str]) -> List[Dict]:
        split_samples = []
        for sid in session_ids:
            if max_samples is not None and len(split_samples) >= max_samples:
                break
            if sid not in session_dirs:
                print(f"[noxiloader] Session {sid} not found in data directory, skipping")
                continue

            session_samples = _build_samples_for_session(
                session_id=sid,
                session_dir=session_dirs[sid],
                au_columns=au_columns,
                max_seq_length=max_seq_length,
                combined_entries=combined_by_session.get(sid),
                summary_entries=summaries_by_session.get(sid),
            )
            split_samples.extend(session_samples)

            if max_samples is not None and len(split_samples) >= max_samples:
                split_samples = split_samples[:max_samples]
                break
        return split_samples

    train_samples = _process_split(train_sessions)
    val_samples = _process_split(val_sessions)
    test_samples = _process_split(test_sessions)

    print(f"[noxiloader] Final sizes: train={len(train_samples)}, "
          f"val={len(val_samples)}, test={len(test_samples)}")

    result = (train_samples, val_samples, test_samples)

    # Save to cache
    try:
        print(f"[noxiloader] Saving to cache: {cache_file}")
        with open(cache_file, "wb") as f:
            pickle.dump(result, f)
    except Exception as e:
        print(f"[noxiloader] Cache save failed: {e}")

    return result


# ============================================================================
# MAIN (for testing)
# ============================================================================

if __name__ == "__main__":
    print("=== NoXi CoT Dataset Loading Demo ===\n")

    train, val, test = load_noxi_cot_splits(
        data_dir="results",
        max_samples=5,
    )

    print(f"\nTrain samples: {len(train)}")
    print(f"Val samples: {len(val)}")
    print(f"Test samples: {len(test)}")

    if train:
        sample = train[0]
        print(f"\nSample from training set:")
        print(f"  Session ID: {sample['session_id']}")
        print(f"  Speaker: {sample['speaker_id']}")
        print(f"  Turn index: {sample['turn_index']}")
        print(f"  Time window: {sample['start_ms']}ms - {sample['end_ms']}ms")
        print(f"  AU columns ({len(sample['au_columns'])}): {sample['au_columns'][:5]}...")
        print(f"  Expert AU vectors keys: {list(sample['expert_au_vectors'].keys())[:5]}...")
        print(f"  Novice AU vectors keys: {list(sample['novice_au_vectors'].keys())[:5]}...")
        first_au = list(sample['expert_au_vectors'].keys())[0]
        print(f"  Expert {first_au} length: {len(sample['expert_au_vectors'][first_au])}")
        print(f"  Answer: {sample['answer'][:150]}...")
